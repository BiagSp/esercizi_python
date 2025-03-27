"""
Creiamo un programma che:

Recupera i dati degli utenti da JSONPlaceholder
Recupera i post scritti da ciascun utente
Recupera e conta i commenti associati a ciascun post
Crea una tabella per ogni utente che mostri i suoi post e il numero di commenti

"""

from prettytable import PrettyTable
import requests
import csv

url = "https://jsonplaceholder.typicode.com/users"

users = requests.get(url).json()


url = "https://jsonplaceholder.typicode.com/posts"

posts = requests.get(url).json()

userid_to_name = {}

#Creiamo il dizionario che mappa ID utente -> username
for u in users:
    userid_to_name[u["id"]] = u["username"]


#organizziamo i post per nome utente
user_posts = {}

for post in posts:
    user_id = post["userId"]
    username = userid_to_name[user_id]

    if username not in user_posts:
        user_posts[username] = []

    user_posts[username].append(post)


#recuperiamo i commenti

url = "https://jsonplaceholder.typicode.com/comments"
comments = requests.get(url).json()

#contiamo i commenti per ogni post

post_comment_count = {}

for comment in comments:
    post_id = comment["postId"]
    if post_id in post_comment_count:
        post_comment_count[post_id] += 1
    else:
        post_comment_count[post_id] = 1



for username, posts in user_posts.items():
    table = PrettyTable() 
    table.field_names = [username, "Comments"]

    for post in posts:
        title = post["title"]
        if len(title) > 30:
            title = title[:20] + "..."
        
        post_id = post["id"]
        comment_count = post_comment_count.get(post_id, 0)
        table.add_row([title, comment_count])

    print(table) 
    print("\n")


#Definiamo il nome del file csv

csv_filename = "user_post_comments.csv"

#apriamo il file in modalità scrittua
with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['Username', 'UserID', 'PostID', 'PostTitle', 'PostBody', 'NumComments'])
    
    for username, posts in user_posts.items():
        for post in posts:
            post_id = post["id"]
            user_id = post["userId"]  # Usa direttamente l'userId dal post
            post_title = post["title"]
            post_body = post["body"]
            comment_count = post_comment_count.get(post_id, 0)
            
            writer.writerow([
                username,
                user_id,
                post_id,
                post_title,
                post_body,
                comment_count
            ])



print(f"File CSV {csv_filename} creato con successo")


import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
# Rimuovi il foglio di default
wb.remove(wb.active)

for username, posts in user_posts.items():
    # Crea un foglio per ogni utente
    ws = wb.create_sheet(title=username)
    
    # Stile dell'intestazione
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Aggiungi un'intestazione personalizzata con il nome utente
    ws.append([f"Posts by {username}"])
    ws.merge_cells('A1:C1')  # Unisci le celle per l'intestazione
    cell = ws['A1']
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')
    
    # Aggiungi intestazioni delle colonne
    ws.append(['Title', 'Content', 'Comments'])
    
    # Applica stile alle intestazioni
    for col in range(1, 4):  # A, B, C
        cell = ws.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
    
    # Aggiungi i dati
    for post in posts:
        post_id = post["id"]
        ws.append([
            post["title"],
            post["body"],
            post_comment_count.get(post_id, 0)
        ])
    
    # Adatta larghezza colonne
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15

# Crea una pagina indice
index = wb.create_sheet(title="Index", index=0)
index.append(["User Index"])
index.append(["Username", "Number of Posts"])

row = 3
for username, posts in user_posts.items():
    index.append([username, len(posts)])
    # Aggiungi hyperlink al foglio utente
    cell = index.cell(row=row, column=1)
    cell.hyperlink = f"#{username}!A1"
    cell.font = Font(color="0000FF", underline="single")
    row += 1

wb.save('user_posts_excel.xlsx')
print("File Excel 'user_posts_excel.xlsx' creato con successo!")