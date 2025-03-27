# excel_exporter.py
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

def export_to_excel(data_dict, metrics_dict, file_path="etf_analysis.xlsx"):
    """
    Esporta i dati in un file Excel strutturato
    
    Args:
        data_dict: Dizionario con simboli come chiavi e DataFrame come valori
        metrics_dict: Dizionario con simboli come chiavi e dizionari di metriche come valori
        file_path: Percorso del file Excel da creare
    """
    # Crea un nuovo file Excel con ExcelWriter
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # 1. Crea il foglio Panoramica
        create_overview_sheet(writer, data_dict, metrics_dict)
        
        # 2. Crea fogli individuali per ogni strumento
        for symbol, df in data_dict.items():
            create_individual_sheet(writer, symbol, df, metrics_dict.get(symbol, {}))
            
        # 3. Crea il foglio di analisi del portafoglio
        create_portfolio_sheet(writer, data_dict, metrics_dict)
        
    # Ora apriamo il file per aggiungere formattazione avanzata e grafici
    wb = load_workbook(file_path)
    
    # Aggiungi formattazione e grafici
    format_workbook(wb, data_dict, metrics_dict)
    
    # Salva il file formattato
    wb.save(file_path)
    print(f"File Excel creato con successo: {file_path}")

def create_overview_sheet(writer, data_dict, metrics_dict):
    """Crea il foglio di panoramica con tabella riepilogativa"""
    # Crea DataFrame di riepilogo
    overview_data = []
    for symbol, metrics in metrics_dict.items():
        row = {'Simbolo': symbol}
        row.update(metrics)
        overview_data.append(row)
    
    overview_df = pd.DataFrame(overview_data)
    
    # Scrivi nel foglio Excel
    overview_df.to_excel(writer, sheet_name='Panoramica', index=False)

def create_individual_sheet(writer, symbol, df, metrics):
    """Crea un foglio individuale per uno strumento"""
    # Prepara i dati
    sheet_df = df.copy()
    
    # Assicurati che la data sia una colonna e non un indice
    if 'date' not in sheet_df.columns and isinstance(sheet_df.index, pd.DatetimeIndex):
        sheet_df = sheet_df.reset_index()
    
    # Formatta la data come stringa se necessario
    if 'date' in sheet_df.columns and pd.api.types.is_datetime64_any_dtype(sheet_df['date']):
        sheet_df['date'] = sheet_df['date'].dt.strftime('%Y-%m-%d')
    
    # Scrivi nel foglio Excel
    sheet_df.to_excel(writer, sheet_name=symbol, index=False)
    
    # Aggiungi le metriche in cima al foglio
    metrics_df = pd.DataFrame([metrics])
    metrics_df.to_excel(writer, sheet_name=symbol, startrow=len(sheet_df) + 3, index=False)

def create_portfolio_sheet(writer, data_dict, metrics_dict):
    """Crea il foglio di analisi del portafoglio"""
    # Crea un semplice esempio di allocazione
    symbols = list(data_dict.keys())
    n_assets = len(symbols)
    
    # Allocazione equamente distribuita come esempio
    allocations = [1/n_assets] * n_assets
    
    portfolio_df = pd.DataFrame({
        'Strumento': symbols,
        'Allocazione (%)': [a * 100 for a in allocations],
        'Ultimo Prezzo': [data_dict[s]['close'].iloc[-1] if 'close' in data_dict[s].columns else 0 
                         for s in symbols],
    })
    
    # Aggiungi metriche rilevanti dal dizionario delle metriche
    for metric in ['Rendimento 1Y (%)', 'Volatilità (%)', 'Sharpe Ratio']:
        if all(metric in metrics_dict[s] for s in symbols):
            portfolio_df[metric] = [metrics_dict[s][metric] for s in symbols]
    
    # Calcola i valori del portafoglio
    portfolio_df['Contributo al Portafoglio (%)'] = portfolio_df['Allocazione (%)'] 
    
    # Scrivi nel foglio Excel
    portfolio_df.to_excel(writer, sheet_name='Portafoglio', index=False)

def format_workbook(wb, data_dict, metrics_dict):
    """Aggiunge formattazione avanzata e grafici al workbook"""
    # Qui implementeremmo la formattazione avanzata con openpyxl
    # e aggiungeremmo i grafici
    pass